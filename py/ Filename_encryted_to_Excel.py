# Collect All the Information from the Filename and Summarize in Excel
# Username can be encryted in to Anonymous ID
import os, shutil
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import datetime

# excel initial
excel_filename = './../Voice_Data_Encryted.xlsx'
wb = load_workbook(filename = excel_filename)
sheet_title = str(datetime.date.today())
ws = wb.create_sheet(title=sheet_title)
ws = wb[sheet_title]
date=datetime.date.today()
formatted_date=date.strftime('%y%m%d')

# files
path = os.getcwd() + "/../Data/"
file_paths = os.listdir(path)
file_paths.sort()
dst_path = "./Data_Encryted/"
excel_line = 1
Username_old = ""

header = ['Username', 'Anonymous_ID','PDPA','Fixed1_Length', 'Fixed1_Time', 'Fixed2_Length', 'Fixed2_Time',
          'Fixed3_Length', 'Fixed3_Time', 'Fixed4_Length', 'Fixed4_Time',
          'Random1_Length', 'Random1_Time', 'Random2_Length', 'Random2_Time']
header_dict = {}

for i, col in enumerate(list(map(chr, range(ord('A'), ord('A')+len(header))))):
    header_dict[header[i]] = col;
    ws[col+str(1)]=header[i]


for i, fp in enumerate(file_paths):
    file_format = os.path.splitext(fp)[-1]
    file_name = os.path.splitext(fp)[0]
    if (file_format == ".wav"):
        # Getting Info From Filename
        Username = str(file_name.split('_')[0])
        TextID = str(file_name.split('_')[1])
        Length = str(file_name.split('_')[2])
        Time = str(file_name.split('_')[3])+':'+str(file_name.split('_')[4])+':'+str(file_name.split('_')[5])
        Date = str(file_name.split('_')[6])+'/'+str(file_name.split('_')[7])+'/'+str(file_name.split('_')[8])
        Time_Date = Time+' '+Date

        if(Username != Username_old):
            excel_line = excel_line + 1
            ws[header_dict['Username']+str(excel_line)]=Username
            # Can change into any kind of Anonymous_ID
            Anonymous_ID = formatted_date + '{:03d}'.format(excel_line-1)
            ws[header_dict['Anonymous_ID']+str(excel_line)]=Anonymous_ID
            ws[header_dict['PDPA']+str(excel_line)]='Agree'
            Username_old = Username

        ws[header_dict[TextID+'_Length']+str(excel_line)]=Length
        ws[header_dict[TextID+'_Time']+str(excel_line)]=Time_Date

        # copy
        shutil.copy2(path+fp, dst_path+fp.replace(Username, Anonymous_ID))


print (excel_line)
wb.save(filename=excel_filename)
